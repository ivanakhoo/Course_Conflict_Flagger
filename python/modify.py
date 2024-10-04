from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import pandas as pd
from openpyxl.worksheet.cell_range import CellRange


def remove_first_row(file_path, output_path):
    # Load the workbook and select the active worksheet
    wb = load_workbook(file_path)
    ws = wb.active
    print(ws)
    
    ws.delete_rows(1) 
    
    # Save the modified workbook to a new file
    wb.save(output_path)

# Work on the delete merged cell method!
def delete_cells(file_path, cell_range):
    # Load the workbook and select the active worksheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Get the boundaries of the range (e.g., G1-K1)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    # List of merged cell ranges
    merged_cells = list(ws.merged_cells.ranges)

    # Function to check if a cell is part of a merged range
    def is_in_merged_range(cell):
        for merged_range in merged_cells:
            min_cr, min_rr, max_cr, max_rr = CellRange(str(merged_range)).bounds
            if min_cr <= cell.column <= max_cr and min_rr <= cell.row <= max_rr:
                return True
        return False

    # Loop through the columns in the range
    for col in range(min_col, max_col + 1):
        for row in range(min_row, ws.max_row):
            cell = ws.cell(row=row, column=col)
            below_cell = ws.cell(row=row + 1, column=col)

            # Skip modification if the cell is part of a merged range
            if is_in_merged_range(cell):
                continue

            # Move the cell value from the row below
            cell.value = below_cell.value

        # Clear the last row's cells in the range
        last_row_cell = ws.cell(row=ws.max_row, column=col)
        if not is_in_merged_range(last_row_cell):
            last_row_cell.value = None

    # Save the modified workbook
    wb.save(file_path)


def insert_cells(file_path):
    # Load the workbook and select the active worksheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Define the row index and column range for insertion (G-K)
    insert_row = 2  # Row number to insert above (G2:K2)
    min_col, max_col = 7, 11  # Column numbers for G (7) to K (11)

    # Shift the rows below the specified range down
    for row in range(ws.max_row, insert_row - 1, -1):  # Start from the bottom
        for col in range(min_col, max_col + 1):
            ws.cell(row=row + 1, column=col).value = ws.cell(row=row, column=col).value

    # Optionally clear the cells in the inserted row (G2:K2)
    for col in range(min_col, max_col + 1):
        ws.cell(row=insert_row, column=col).value = None

    # Save the modified workbook to a new file
    wb.save(file_path)


def rename_column(file_path):
    # Load the workbook and select the active worksheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Find the column with header "Course prefix & number" and rename it
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == "Course prefix & number":
            col[0].value = "Courses"
            break

    # Save the modified workbook
    wb.save(file_path)

## PANDAS EQUIVALENTS


# def remove_headers(file_path, output_path):
#     # Reset the column headers to default integers (like 0, 1, 2, ...)
#     df = pd.read_excel(file_path)
#     df.columns = range(df.shape[1])
#     print(df)
    
#     return df.to_excel(file_path)

# def remove_first_row(file_path):
#     # Read the Excel file into a DataFrame
#     df = pd.read_excel(file_path)

    
#     print("original")
#     print(df)
    
#     # Remove the first row
#     df = df.iloc[1:, :]

#     print()
#     print("remove first row")
#     print()
#     print(df)

#     # Save the modified DataFrame to a new Excel file
#     df.to_excel(file_path, index=False)

# def delete_cells(file_path, cell_range):
#     # Read the Excel file into a DataFrame
#     df = pd.read_excel(file_path)

#     print()
#     print("delete cells")
#     print()

    
    
#     # Define the columns and rows to delete
#     start_col, end_col = cell_range.split(':')
    
#     # For simplicity, let's drop the specified range by setting values to NaN
#     df.loc[:, start_col:end_col] = None

#     print(df)

#     # Save the modified DataFrame to the file
#     df.to_excel(file_path, index=False)

# def insert_cells(file_path, output_path):
#     # Read the Excel file into a DataFrame
#     df = pd.read_excel(file_path)

#     print()
#     print("insert cells")
#     print()
    
#     # Define the row index and column range for insertion (G-K)
#     insert_row = 2  # Insert above this row (G2:K2)
    
#     # Insert an empty row with NaN in the specified column range
#     new_row = pd.Series([None] * len(df.columns), index=df.columns)
    
#     # Insert the new row at the specified position
#     df = pd.concat([df.iloc[:insert_row-1], pd.DataFrame([new_row]), df.iloc[insert_row-1:]]).reset_index(drop=True)
    
#     print(df)

#     # Save the modified DataFrame to the file
#     df.to_excel(output_path, index=False)

# def rename_column(file_path, output_path):
#     # Read the Excel file into a DataFrame
#     df = pd.read_excel(file_path)

#     print()
#     print("rename column")
#     print()

   
    
#     # Rename the column if it exists
#     df = df.rename(columns={"Course prefix & number": "Courses"})

#     print(df)
    
#     # Save the modified DataFrame to the file
#     df.to_excel(output_path, index=False)

