import re
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# WHAT TO INPUT:
# python3 Flagger.py [Course Schedule] [Course Conflict Survey]


# Merge the days columns into 'Days'.
def daysMerger(original):
    # Merge five columns into one with a delimiter (for example, comma)
    original['Days'] = original[['M', 'T', 'W', 'H', 'F']].apply(lambda row: ','.join(str(val) for val in row if pd.notna(val)), axis=1)

    # Optionally, you can drop the original columns if needed
    original.drop(columns=['M', 'T', 'W', 'H', 'F'], inplace=True)

    return original


# Compilation of the courses in each course conflict group (array) stored in a master array.
def parse_master_course_conflict(conflicts):
    master_course_conflict_array = []

    for column in conflicts.columns:
        if column.lower().startswith("course group"):
            courses = conflicts[column].apply(lambda x: x.split(", ") if pd.notna(x) else []).tolist()
            master_course_conflict_array.extend(courses)

    return master_course_conflict_array


# Deletes empty arrays from the master array.
def deleteEmpty(master_course_conflict_array):
    master_course_conflict_array[:] = [subarray for subarray in master_course_conflict_array if subarray]


# Takes a single course conflict group from the master_course_conflict_array and the original
# Purpose: Checks to see which courses are in the original, creating a dataframe from it.
def course_conflict_dataframe_maker(course_conflict_group, original):
    matching_rows = pd.DataFrame()
    for course_names in course_conflict_group:
        if course_names:  # Check if the array is not empty
            for course_name in course_names:
                if pd.notna(course_name):
                    mask = original['Courses'].str.contains(str(course_name), na=False)
                    matching_rows = pd.concat([matching_rows, original[mask]], ignore_index=True)
    return matching_rows


# Creates an array of dataframes based off of the course conflict groups and the course schedule. 
# There is one dataframe for each course conflict group.
def master_dataframe_maker(master_course_conflict_array, original):
    dfs = []  # List to store DataFrames
    for course_group_array in master_course_conflict_array:
        matching_rows = course_conflict_dataframe_maker([course_group_array], original)
        remove_commas(matching_rows)
        day_checker(matching_rows)
        dfs.append(matching_rows)
    return dfs


# Removes commas from the 'Days' column.
def remove_commas(df):
    df['Days'] = df['Days'].str.replace(',', '')


# Given a data frame, returns a set of matching pairs
def day_checker(df):
    # Set to store pairs of indices that have been matched
    matched_pairs = []
    
    # Iterate over each row in the 'Days' column
    for i, row in df.iterrows():
        days_set_i = set(row['Days'])
        course_name_i = row['Courses']
        # Iterate over each other row in the 'Days' column
        for j, other_row in df.iterrows():
            days_set_j = set(other_row['Days'])
            course_name_j = other_row['Courses']
            # Skip self-comparison and pairs that have already been matched
            if i != j and (course_name_i != course_name_j) and (i, j) not in matched_pairs and (j, i) not in matched_pairs:
                # Check if the 'Days' values match
                if days_set_i & days_set_j:
                    course1 = row['Courses']
                    course2 = other_row['Courses']
                    # print(f"{course1} matches with {course2}")
                    # Add the pair of indices to the set of matched pairs
                    matched_pairs.append((i, j))
    return matched_pairs


# Converts the BegTime and EndTime column values to an integer without ':'
def convertTime(df):
    df['BegTime'] = df['BegTime'].fillna(0)
    df['EndTime'] = df['EndTime'].fillna(0)

    df['BegTime'] = df['BegTime'].astype(str).str.replace(':', '').astype(int)
    df['EndTime'] = df['EndTime'].astype(str).str.replace(':', '').astype(int)

    return df


# Deletes blank rows
def delete_blank_rows(df):
    
    # Drop rows containing NaN values
    df.dropna(axis=0, how='all', inplace=True)
    return df


# Check if the times overlap at all given two start and two end times
def is_time_overlap(start_a, end_a, start_b, end_b):
    return ((max(start_a, start_b) < min(end_a, end_b)) or (start_a == start_b and end_a == end_b))


# Passed to this method from day_checker and master_dataframe_maker
# Given a matching pair and the corresponding data frame with the course conflict group -> returns matching_pair if overlap, nothing if not
def final_conflicts(matching_pair, df):
    a, b = matching_pair
    start_a = df.at[a, 'BegTime']
    start_b = df.at[b, 'BegTime']
    end_a = df.at[a, 'EndTime']
    end_b = df.at[b, 'EndTime']

    a = df.at[a, 'Courses']
    b = df.at[b, 'Courses']
    matching_pair = (a, b)

    if is_time_overlap(start_a, end_a, start_b, end_b):
        return matching_pair
    else:
        return
    

# Clears out all None values
def none_clearer(matching_pairs):
    matching_pairs = [x for x in matching_pairs if x is not None]
    return matching_pairs


# Create a list of pairs that are conflicting/overlapping!
# listofDfs - master_dataframe_maker
def compiler(listOfDFs):
    compiledList = []
    for df in listOfDFs:
        day_checker_output = day_checker(df)
        for matching_pair in day_checker_output:
            compiledList.append(final_conflicts(matching_pair, df))

    compiledList = none_clearer(compiledList)
    return compiledList


# Find row indices of the pairs
def row_index_finder(compiledList, original):
    newCompiledList = []
    for pair in compiledList:
        a, b = pair
        index_a = original[original['Courses'] == a].index.tolist()
        index_b = original[original['Courses'] == b].index.tolist()
        pair = (index_a, index_b)
        newCompiledList.append(pair)

    return newCompiledList


# Highlights the rows of the conflicting course pairs
def highlighter(compiledList, file_path, new_file_path):
    # Load the Excel file
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define a fill color for highlighting
    highlight_color = "FFFF00"  # Yellow fill color
    
    # Iterate over each pair of indices and perform the highlighting/writing
    for pair in compiledList:
        idx_a, idx_b = pair
        
        # Adjust indices to 1-based for Excel (row_idx)
        row_idx_a = idx_a[0] + 2  # DataFrame indices start from 0, Excel rows from 1
        row_idx_b = idx_b[0] + 2  # Adjust for 1-based index
        
        #This is probably where I need to change stuff

        # Highlight and write for course A
        highlight_and_write(ws, row_idx_a, row_idx_b, highlight_color)
        
        # Highlight and write for course B
        highlight_and_write(ws, row_idx_b, row_idx_a, highlight_color)
    
    # Save the modified Excel file
    wb.save(new_file_path)



# Helper function to highlight and write conflicting row indices
def highlight_and_write(ws, row_idx, conflicting_row_idx, highlight_color):
    # Highlight the row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")
    
    # Write the conflicting row information in the first column
    current_value = ws.cell(row=row_idx, column=1).value
    if current_value:
        # If the cell is not empty, append the conflicting row
        ws.cell(row=row_idx, column=1, value=f"{current_value}, {conflicting_row_idx}")
    else:
        # If the cell is empty, write the conflicting row information
        ws.cell(row=row_idx, column=1, value=f"Conflicting with row indices: {conflicting_row_idx}")



# Check if all courses in a dataframe conflict, classifying as hard, medium, or soft.
def classify_conflict(df):

    total_rows = len(df)
    matched_pairs = day_checker(df)  # Ensure df is a valid DataFrame here
    
    # Hard conflict: All courses in the dataframe conflict.
    if len(matched_pairs) == (total_rows * (total_rows - 1)) // 2:
        return "hard", matched_pairs
    
    # Medium conflict: Some courses conflict, but there are alternative sections.
    if len(matched_pairs) > 0:
        return "medium", matched_pairs
    
    # No conflicts: Return None or some indicator of no conflicts.
    return "none", []



def main():

    # Course Schedule is read in.
    courseSchedule = sys.argv[1]


    # Course Conflicts Survey is read in.
    courseConflict = sys.argv[2]

    process(courseSchedule, courseConflict)




def process(courseSchedule, courseConflict, new_file_path):

    # Convert courseConflicts into a dataframe.
    courseConflicts = pd.read_excel(courseConflict)


    # Convert courseSchedule into a dataframe.
    original = pd.read_excel(courseSchedule)


    # Delete blank rows from the excel file.
    delete_blank_rows(original)


    # Convert the time columns to integers without the ":".
    convertTime(original)


    # Merge the days columns into 'Days'.
    daysMerger(original)


    # Compilation of the courses in each course conflict group (array) stored in a master array.
    conflictsArray = parse_master_course_conflict(courseConflicts)


    # Deletes the empty arrays in the master array (if any of the courses in the course conflict groups are not in the course schedule).
    deleteEmpty(conflictsArray)
   

    # Creates an array of dataframes based off of the course conflict groups and the course schedule. 
    # There is one dataframe for each course conflict group.
    conflictRows = master_dataframe_maker(conflictsArray, original)


    # Compiles a list of pairs of course names that conflict in day and time from the course conflict groups.
    compiledList = compiler(conflictRows)


    # Returns the row indices of the pairs in the compiledList
    newCompiledList = row_index_finder(compiledList, original)

    # Highlights the rows of the conflicting course pairs
    highlighter(newCompiledList, courseSchedule, new_file_path)


    return new_file_path


# Runs the program.
if __name__ == "__main__":
    main()


